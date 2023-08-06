keep_going = True
word_list = []
while keep_going:
    new_word = input("Enter the word you want and continue or ENTER to quit:")
    if new_word != "":
        word_list.append(new_word)
    keep_going = bool(new_word)
import os
import openpyxl
import requests
import json
path = r"D:\coding"
os.chdir(path)

workbook = openpyxl.Workbook()
sheet = workbook.active  
keep_going = True
for ans in word_list:
    response_api = requests.get('https://api.dictionaryapi.dev/api/v2/entries/en/%s'%ans)
#print(response_api.content)
    word_def = response_api.content


# some JSON:
    x = word_def
# parse x:
    y = json.loads(x)
    z = json.dumps(y)

# the result is a Python dictionary:
#print(y)
#print(y[0]["meanings"][0]["partOfSpeech"])
#print(y[0]["meanings"])
#print(len(["definition"]))
    print("--------------------------------Processing----------------------------------")
 

   

    #sheet.title = '1st sheet'
    #sheet['A1'].value = "word"
    #sheet['E1'].value = "meanings"
    #sheet['J1'].value = "part of speech"
    #sheet['A2'].value = "%s"%ans
#print(y[0]["meanings"][0]["definitions"][0]["example"])
    while keep_going:
        max_rows = sheet.max_row
        sheet['A%d'%(max_rows + 2)] = ans   
        sheet['A%d'%(max_rows + 2)].font = openpyxl.styles.Font(name = "Times New Romen",size = 12, bold = True, italic = False)   
        for i in range(len(y[0]["meanings"])):
            #print(y[0]["meanings"][x]["definitions"])
            #print(pd.DataFrame(y[0]["meanings"][i]["definitions"]))
            #max_rows = sheet.max_row
            #defintion = (y[0]["meanings"][i]["definitions"][0]["definition"])
            #part_of_speech = (y[0]["meanings"][i]["partOfSpeech"])
            #sheet['A%d'% (max_rows+3)] = defintion
            #sheet['A%d'% (max_rows+3)].font = openpyxl.styles.Font(name = "Times New Romen",size = 12, bold = False, italic = True) 
            #sheet['D%d'% (max_rows+2)] = part_of_speech
            #sheet['D%d'% (max_rows+2)].font = openpyxl.styles.Font(name = "Times New Romen",size = 12, bold = False, italic = False) 
         #   #workbook.save('1.xlsx')
            for n in range(len(y[0]["meanings"][i]["definitions"])):
                max_rows = sheet.max_row
                defintion = (y[0]["meanings"][i]["definitions"][n]["definition"])
                part_of_speech = (y[0]["meanings"][i]["partOfSpeech"])
                sheet['A%d'% (max_rows+3)] = defintion
                sheet['A%d'% (max_rows+3)].font = openpyxl.styles.Font(name = "Times New Romen",size = 12, bold = False, italic = True) 
                sheet['D%d'% (max_rows+2)] = part_of_speech
                sheet['D%d'% (max_rows+2)].font = openpyxl.styles.Font(name = "Times New Romen",size = 12, bold = False, italic = False) 
                workbook.save('wordlist.xlsx')
        #
        break
sheet.delete_rows(idx=1, amount=2)
workbook.save('wordlist.xlsx')
print("----------------------process finished successfully------------------------------")